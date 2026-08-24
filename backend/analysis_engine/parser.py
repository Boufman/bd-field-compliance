"""
Field Compliance – Excel Service Call Intake parser.
Parses BD / WA Health Service Call Intake workbook into structured records.
"""
from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Optional
from datetime import datetime, date
import openpyxl


COL_MAP = {
    "at_risk": 1,
    "date_submitted": 2,
    "work_order": 3,
    "product_classification": 4,
    "asset_number": 5,
    "job_description": 6,
    "additional_info": 7,
    "completed_by": 8,
    "rss_manager": 9,
    "department": 10,
    "time_submitted": 11,
    "date_start": 12,
    "between_8_and_8": 13,
    "time_start": 14,
    "completed_date": 15,
    "completed_time": 16,
    "callout": 17,
    "completion_comments": 18,
    "analysis": 19,
    "part_dispensing": 20,
    "qty_used": 21,
    "part_infusion": 22,
    "severity": 23,
    "completion_code": 24,
    "duration_case_to_start_str": 25,
    "duration_case_to_start_min": 26,
    "duration_tech_str": 27,
    "duration_tech_min": 28,
    "total_time_resolved": 29,
    "breach": 30,
    "within_date_range": 31,
    "week_number": 32,
    "site": 33,
    "dispensing_or_infusion": 34,
    "tsc_downtime": 35,
    "cubie": 36,
    "service_num": 37,
    "archive": 38,
    "saturday_or_sunday": 39,
    "weekend": 40,
}


def _cell_val(ws, row: int, col: int) -> Any:
    v = ws.cell(row=row, column=col).value
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        if v == "":
            return None
    return v


def _parse_date(val) -> Optional[date]:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(val[:10], fmt).date()
            except ValueError:
                continue
    return None


def _parse_bool(val) -> bool:
    if val is None:
        return False
    if isinstance(val, bool):
        return val
    return str(val).strip().upper() in ("TRUE", "YES", "1", "Y", "T")


def _safe_str(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _clean_work_order(val) -> str:
    """Force clean integer string – no decimals."""
    if val is None:
        return ""
    try:
        return str(int(float(val)))
    except (TypeError, ValueError):
        return str(val).strip()


def parse_excel(source: BytesIO | Path | str, original_filename: str = "intake.xlsx") -> Dict[str, Any]:
    if isinstance(source, (str, Path)):
        wb = openpyxl.load_workbook(source, data_only=True)
    else:
        source.seek(0)
        wb = openpyxl.load_workbook(source, data_only=True)

    ws = None
    for name in wb.sheetnames:
        if "service call" in name.lower() or "intake" in name.lower():
            ws = wb[name]
            break
    if ws is None:
        ws = wb.active

    records: List[Dict[str, Any]] = []
    for row in range(2, ws.max_row + 1):
        wo = _cell_val(ws, row, COL_MAP["work_order"])
        if wo is None:
            continue

        rec = {
            "work_order": _clean_work_order(wo),
            "date_submitted": _parse_date(_cell_val(ws, row, COL_MAP["date_submitted"])),
            "time_submitted": _safe_str(_cell_val(ws, row, COL_MAP["time_submitted"])),
            "date_start": _parse_date(_cell_val(ws, row, COL_MAP["date_start"])),
            "time_start": _safe_str(_cell_val(ws, row, COL_MAP["time_start"])),
            "date_completed": _parse_date(_cell_val(ws, row, COL_MAP["completed_date"])),
            "time_completed": _safe_str(_cell_val(ws, row, COL_MAP["completed_time"])),
            "product_classification": _safe_str(_cell_val(ws, row, COL_MAP["product_classification"])),
            "asset_number": _safe_str(_cell_val(ws, row, COL_MAP["asset_number"])),
            "job_description": _safe_str(_cell_val(ws, row, COL_MAP["job_description"])),
            "completion_comments": _safe_str(_cell_val(ws, row, COL_MAP["completion_comments"])),
            "department": _safe_str(_cell_val(ws, row, COL_MAP["department"])),
            "completed_by": _safe_str(_cell_val(ws, row, COL_MAP["completed_by"])),
            "severity": _safe_str(_cell_val(ws, row, COL_MAP["severity"])) or "Non-Urgent",
            "completion_code": _safe_str(_cell_val(ws, row, COL_MAP["completion_code"])) or "Other",
            "between_8_and_8": _parse_bool(_cell_val(ws, row, COL_MAP["between_8_and_8"])),
            "cubie_replaced": _parse_bool(_cell_val(ws, row, COL_MAP["cubie"])),
            "saturday_or_sunday": _safe_str(_cell_val(ws, row, COL_MAP["saturday_or_sunday"])),
            "weekend": _parse_bool(_cell_val(ws, row, COL_MAP["weekend"])),
            "breach": _safe_str(_cell_val(ws, row, COL_MAP["breach"])),
            "week_number": _cell_val(ws, row, COL_MAP["week_number"]),
            "site": _safe_str(_cell_val(ws, row, COL_MAP["site"])) or "Fiona Stanley Hospital",
            "service_num": _safe_str(_cell_val(ws, row, COL_MAP["service_num"])),
            "duration_case_to_start_min": _cell_val(ws, row, COL_MAP["duration_case_to_start_min"]),
            "duration_tech_min": _cell_val(ws, row, COL_MAP["duration_tech_min"]),
            "total_time_resolved": _cell_val(ws, row, COL_MAP["total_time_resolved"]),
            "part_used": _safe_str(_cell_val(ws, row, COL_MAP["part_dispensing"]))
                        or _safe_str(_cell_val(ws, row, COL_MAP["part_infusion"])),
            "qty_used": _cell_val(ws, row, COL_MAP["qty_used"]),
        }
        records.append(rec)

    records.sort(key=lambda r: r["date_submitted"] or date.min, reverse=True)

    return {
        "filename": original_filename,
        "record_count": len(records),
        "records": records,
        "parsed_at": datetime.utcnow().isoformat() + "Z",
    }


def classify_with_preview(source: BytesIO, original_filename: str = "intake.xlsx", max_points: int = 50) -> Dict[str, Any]:
    data = parse_excel(source, original_filename)
    return {
        "filename": data["filename"],
        "record_count": data["record_count"],
        "preview_count": len(data["records"][:max_points]),
        "preview": data["records"][:max_points],
        "parsed_at": data["parsed_at"],
    }